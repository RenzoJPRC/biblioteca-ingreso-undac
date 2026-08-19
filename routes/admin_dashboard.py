from flask import Blueprint, render_template, request, Response, session, send_file
from utils.queries_dashboard import obtener_datos_dashboard, obtener_registros_csv
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import time
import io
import pandas as pd

# El Blueprint para el dashboard y la raíz de /admin
admin_dashboard_bp = Blueprint('admin_dashboard', __name__, url_prefix='/admin')

@admin_dashboard_bp.route('/')
def admin_dashboard():
    f_inicio = request.args.get('inicio')
    f_fin = request.args.get('fin')
    hora_inicio = request.args.get('hora_inicio')
    hora_fin = request.args.get('hora_fin')
    
    sede_filtro = session.get('admin_sede') if session.get('admin_rol') == 'Supervisor' else None
    dash_data = obtener_datos_dashboard(f_inicio, f_fin, sede_filtro, hora_inicio, hora_fin)

    return render_template('admin_dashboard.html', 
                           total_hoy=dash_data['total_hoy'], 
                           total_alumnos=dash_data['total_alumnos'],
                           total_visitantes=dash_data['total_visitantes'], 
                           total_egresados=dash_data['total_egresados'], 
                           total_personal=dash_data['total_personal'],
                           total_docentes=dash_data.get('total_docentes', 0),
                           pisos=dash_data['pisos'], 
                           salas=dash_data.get('salas', {}),
                           sedes=dash_data['sedes'],
                           ultimos=dash_data['ultimos_crudos'], 
                           labels_horas=dash_data['chart_horas_labels'], 
                           data_horas=dash_data['chart_horas_values'], 
                           labels_escuelas=dash_data['chart_escuelas_labels'], 
                           data_escuelas=dash_data['chart_escuelas_values'], 
                           filtro_label=dash_data['filtro_label'], 
                           params_inicio=f_inicio or '', 
                           params_fin=f_fin or '',
                           params_hora_inicio=hora_inicio or '',
                           params_hora_fin=hora_fin or '')

@admin_dashboard_bp.route('/api/dashboard_stream')
def api_dashboard_stream():
    sede_filtro = session.get('admin_sede') if session.get('admin_rol') == 'Supervisor' else None
    
    def generate():
        while True:
            try:
                dash_data = obtener_datos_dashboard(None, None, sede_filtro, None, None)
                payload = {
                    'total_hoy': dash_data['total_hoy'],
                    'total_alumnos': dash_data['total_alumnos'],
                    'total_visitantes': dash_data['total_visitantes'],
                    'total_egresados': dash_data['total_egresados'],
                    'total_personal': dash_data['total_personal'],
                    'total_docentes': dash_data.get('total_docentes', 0),
                    'pisos': dash_data['pisos'],
                    'salas': dash_data.get('salas', {}),
                    'sedes': dash_data['sedes'],
                    'ultimos': dash_data['ultimos']
                }
                yield f"data: {json.dumps(payload)}\n\n"
            except Exception as e:
                print("SSE Error:", e)
            time.sleep(5)
            
    return Response(generate(), mimetype='text/event-stream')

@admin_dashboard_bp.route('/exportar_ingresos_excel')
def exportar_ingresos_excel():
    f_inicio = request.args.get('inicio')
    f_fin = request.args.get('fin')
    hora_inicio = request.args.get('hora_inicio')
    hora_fin = request.args.get('hora_fin')

    sede_filtro = session.get('admin_sede') if session.get('admin_rol') == 'Supervisor' else None

    registros = obtener_registros_csv(f_inicio, f_fin, sede_filtro, hora_inicio, hora_fin)

    datos_limpios = [tuple(r) for r in registros]

    columnas = [
        'ID',
        'Usuario',
        'DNI',
        'Código Matrícula',
        'Perfil',
        'Sede',
        'Piso',
        'Sala',
        'Turno',
        'Fecha',
        'Hora',
        'Facultad / Área',
        'Escuela / Institución / Oficina'
    ]

    df = pd.DataFrame(datos_limpios, columns=columnas)

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Ingresos')

        ws = writer.sheets['Ingresos']

        # Estilo del encabezado
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style='thin', color='D9E2F3'),
            right=Side(style='thin', color='D9E2F3'),
            top=Side(style='thin', color='D9E2F3'),
            bottom=Side(style='thin', color='D9E2F3')
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Bordes y alineación del contenido
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Congelar encabezado
        ws.freeze_panes = "A2"

        # Autofiltro
        ws.auto_filter.ref = ws.dimensions

        # Ajustar ancho de columnas
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 3, 45)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Altura del encabezado
        ws.row_dimensions[1].height = 22

    output.seek(0)

    if f_inicio and f_fin:
        filename = f"Reporte_Ingresos_{f_inicio}_al_{f_fin}.xlsx"
    else:
        filename = "Reporte_Ingresos_Hoy.xlsx"

    return send_file(
        output,
        download_name=filename,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@admin_dashboard_bp.route('/imprimir_reporte')
def imprimir_reporte():
    f_inicio = request.args.get('inicio')
    f_fin = request.args.get('fin')
    hora_inicio = request.args.get('hora_inicio')
    hora_fin = request.args.get('hora_fin')

    sede_filtro = session.get('admin_sede') if session.get('admin_rol') == 'Supervisor' else None

    registros = obtener_registros_csv(f_inicio, f_fin, sede_filtro, hora_inicio, hora_fin)

    # Determinar título
    titulo = "Reporte General de Ingresos"
    if f_inicio and f_fin:
        if f_inicio == f_fin:
            titulo = f"Reporte de Ingresos del {f_inicio}"
        else:
            titulo = f"Reporte de Ingresos desde {f_inicio} hasta {f_fin}"
    else:
        titulo = "Reporte de Ingresos de Hoy"
        
    if hora_inicio and hora_fin:
        titulo += f" (De {hora_inicio} a {hora_fin})"

    return render_template('admin_reporte_print.html', registros=registros, titulo=titulo)
